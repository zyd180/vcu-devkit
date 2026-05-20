"""Tests for core.generators.report_generator — Excel report generation."""

import pytest
from pathlib import Path

from core.parsers.dbc_parser import DBCData, MessageDef, SignalDef
from core.diff.dbc_diff import DBCDiffResult, MessageDiff, SignalDiff, DiffType
from core.generators.report_generator import ReportGenerator


def _make_signal(name, start_bit=0, bit_length=8, byte_order="little_endian",
                 value_type="unsigned", factor=1.0, offset=0.0, minimum=0.0,
                 maximum=255.0, unit="", comment="", receivers=None):
    return SignalDef(
        name=name, start_bit=start_bit, bit_length=bit_length,
        byte_order=byte_order, value_type=value_type,
        factor=factor, offset=offset, minimum=minimum, maximum=maximum,
        unit=unit, comment=comment, receivers=receivers or [],
        value_descriptions={}, mux=None,
    )


def _make_message(name, msg_id=0x100, dlc=8, sender="VCU", signals=None):
    return MessageDef(
        id=msg_id, name=name, dlc=dlc, sender=sender,
        signals=signals or [], comment="", is_extended=False,
    )


def _make_dbc(messages):
    return DBCData(
        version="v1", messages=messages, nodes=[],
        value_tables={}, comments={}, attributes={}, source_path="<test>",
    )


class TestReportGenerator:

    def setup_method(self):
        self.gen = ReportGenerator()

    def test_signal_matrix(self, tmp_path):
        data = _make_dbc([
            _make_message("M1", signals=[
                _make_signal("S1", unit="V", receivers=["BMS"]),
                _make_signal("S2", factor=0.1, offset=-500),
            ]),
        ])
        out = tmp_path / "matrix.xlsx"
        self.gen.generate_signal_matrix(data, out)
        assert out.exists()
        assert out.stat().st_size > 0

    def test_signal_matrix_empty(self, tmp_path):
        data = _make_dbc([])
        out = tmp_path / "empty.xlsx"
        self.gen.generate_signal_matrix(data, out)
        assert out.exists()

    def test_diff_report(self, tmp_path):
        diff = DBCDiffResult(
            old_version="v1", new_version="v2",
            message_diffs=[
                MessageDiff(
                    message_name="M1", diff_type=DiffType.MODIFIED, id=0x100,
                    signal_diffs=[
                        SignalDiff(
                            signal_name="S1",
                            diff_type=DiffType.MODIFIED,
                            changes={"factor": (1.0, 0.5)},
                            message_name="M1",
                        ),
                    ],
                ),
                MessageDiff(
                    message_name="M2", diff_type=DiffType.ADDED, id=0x200,
                    signal_diffs=[
                        SignalDiff(signal_name="S2", diff_type=DiffType.ADDED, message_name="M2"),
                    ],
                ),
            ],
            summary={"messages_added": 1, "messages_modified": 1},
        )
        out = tmp_path / "diff_report.xlsx"
        self.gen.generate_diff_report(diff, out)
        assert out.exists()

    # ── P0: Verify Excel content ───────────────────────────────────────────

    def test_signal_matrix_sheet_name_and_headers(self, tmp_path):
        from openpyxl import load_workbook
        data = _make_dbc([_make_message("M1", signals=[_make_signal("S1")])])
        out = tmp_path / "matrix.xlsx"
        self.gen.generate_signal_matrix(data, out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.title == "Signal Matrix"
        headers = [ws.cell(row=1, column=c).value for c in range(1, 17)]
        assert headers[0] == "报文名"
        assert headers[4] == "信号名"
        assert headers[14] == "接收方"
        assert headers[15] == "值描述"
        wb.close()

    def test_signal_matrix_data_values(self, tmp_path):
        from openpyxl import load_workbook
        data = _make_dbc([_make_message("M1", msg_id=0x123, dlc=8, sender="VCU",
                          signals=[_make_signal("EngSpeed", unit="rpm", receivers=["BMS"])])])
        out = tmp_path / "matrix.xlsx"
        self.gen.generate_signal_matrix(data, out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "M1"
        assert ws.cell(row=2, column=2).value == "0x123"
        assert ws.cell(row=2, column=3).value == 8
        assert ws.cell(row=2, column=4).value == "VCU"
        assert ws.cell(row=2, column=5).value == "EngSpeed"
        assert ws.cell(row=2, column=14).value == "rpm"
        assert ws.cell(row=2, column=15).value == "BMS"
        wb.close()

    def test_signal_matrix_freeze_panes(self, tmp_path):
        from openpyxl import load_workbook
        data = _make_dbc([_make_message("M1", signals=[_make_signal("S1")])])
        out = tmp_path / "matrix.xlsx"
        self.gen.generate_signal_matrix(data, out)
        wb = load_workbook(out)
        assert wb.active.freeze_panes == "A2"
        wb.close()

    def test_message_summary_content(self, tmp_path):
        from openpyxl import load_workbook
        data = _make_dbc([
            _make_message("M1", msg_id=0x100, dlc=8, sender="VCU",
                          signals=[_make_signal("S1"), _make_signal("S2")]),
            _make_message("M2", msg_id=0x200, dlc=4, sender="BMS",
                          signals=[_make_signal("S3")]),
        ])
        out = tmp_path / "summary.xlsx"
        self.gen.generate_message_summary(data, out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.title == "Message Summary"
        # Headers
        headers = [ws.cell(row=1, column=c).value for c in range(1, 7)]
        assert headers == ["报文名", "CAN ID", "DLC (bytes)", "发送方", "信号数量", "备注"]
        # Row 1: M1
        assert ws.cell(row=2, column=1).value == "M1"
        assert ws.cell(row=2, column=2).value == "0x100"
        assert ws.cell(row=2, column=5).value == 2
        # Row 2: M2
        assert ws.cell(row=3, column=1).value == "M2"
        assert ws.cell(row=3, column=2).value == "0x200"
        assert ws.cell(row=3, column=5).value == 1
        wb.close()

    def test_diff_report_sheets_exist(self, tmp_path):
        from openpyxl import load_workbook
        diff = DBCDiffResult(
            old_version="v1", new_version="v2",
            message_diffs=[
                MessageDiff(
                    message_name="M1", diff_type=DiffType.ADDED, id=0x100,
                    signal_diffs=[
                        SignalDiff(signal_name="S1", diff_type=DiffType.ADDED, message_name="M1"),
                    ],
                ),
            ],
            summary={"messages_added": 1},
        )
        out = tmp_path / "diff.xlsx"
        self.gen.generate_diff_report(diff, out)
        wb = load_workbook(out)
        assert "变更汇总" in wb.sheetnames
        assert "变更详情" in wb.sheetnames
        wb.close()

    def test_diff_report_summary_content(self, tmp_path):
        from openpyxl import load_workbook
        diff = DBCDiffResult(
            old_version="v1", new_version="v2",
            message_diffs=[],
            summary={"messages_added": 2, "signals_removed": 3},
        )
        out = tmp_path / "diff.xlsx"
        self.gen.generate_diff_report(diff, out)
        wb = load_workbook(out)
        ws = wb["变更汇总"]
        assert ws.cell(row=2, column=1).value == "旧版本: v1"
        assert ws.cell(row=3, column=1).value == "新版本: v2"
        assert ws.cell(row=6, column=1).value == "新增报文"
        assert ws.cell(row=6, column=2).value == 2
        assert ws.cell(row=7, column=1).value == "删除信号"
        assert ws.cell(row=7, column=2).value == 3
        wb.close()

    def test_diff_report_detail_with_changes(self, tmp_path):
        from openpyxl import load_workbook
        diff = DBCDiffResult(
            old_version="v1", new_version="v2",
            message_diffs=[
                MessageDiff(
                    message_name="M1", diff_type=DiffType.MODIFIED, id=0x100,
                    signal_diffs=[
                        SignalDiff(
                            signal_name="S1", diff_type=DiffType.MODIFIED,
                            changes={"factor": (1.0, 0.5)},
                            message_name="M1",
                        ),
                    ],
                ),
            ],
            summary={"messages_modified": 1},
        )
        out = tmp_path / "diff.xlsx"
        self.gen.generate_diff_report(diff, out)
        wb = load_workbook(out)
        ws = wb["变更详情"]
        # Header row
        headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
        assert headers == ["操作", "报文", "CAN ID", "信号", "变更字段", "旧值", "新值"]
        # Data row
        assert ws.cell(row=2, column=1).value == "修改"
        assert ws.cell(row=2, column=2).value == "M1"
        assert ws.cell(row=2, column=5).value == "factor"
        assert ws.cell(row=2, column=6).value == "1.0"
        assert ws.cell(row=2, column=7).value == "0.5"
        # Fill color (yellow for MODIFIED)
        fill = ws.cell(row=2, column=1).fill
        assert fill.start_color.rgb == "00FFF5B1"
        wb.close()

    def test_diff_report_signal_added_without_changes(self, tmp_path):
        from openpyxl import load_workbook
        diff = DBCDiffResult(
            old_version="v1", new_version="v2",
            message_diffs=[
                MessageDiff(
                    message_name="M2", diff_type=DiffType.ADDED, id=0x200,
                    signal_diffs=[
                        SignalDiff(signal_name="NewSig", diff_type=DiffType.ADDED, message_name="M2"),
                    ],
                ),
            ],
            summary={"messages_added": 1},
        )
        out = tmp_path / "diff.xlsx"
        self.gen.generate_diff_report(diff, out)
        wb = load_workbook(out)
        ws = wb["变更详情"]
        assert ws.cell(row=2, column=1).value == "新增"
        assert ws.cell(row=2, column=2).value == "M2"
        assert ws.cell(row=2, column=4).value == "NewSig"
        # Green fill for ADDED
        fill = ws.cell(row=2, column=1).fill
        assert fill.start_color.rgb == "00E6FFED"
        wb.close()

    def test_diff_report_message_level_diff(self, tmp_path):
        """Message diff without signal_diffs — message-level row."""
        from openpyxl import load_workbook
        diff = DBCDiffResult(
            old_version="v1", new_version="v2",
            message_diffs=[
                MessageDiff(
                    message_name="OldMsg", diff_type=DiffType.REMOVED, id=0x300,
                    signal_diffs=[],
                ),
            ],
            summary={"messages_removed": 1},
        )
        out = tmp_path / "diff.xlsx"
        self.gen.generate_diff_report(diff, out)
        wb = load_workbook(out)
        ws = wb["变更详情"]
        assert ws.cell(row=2, column=1).value == "删除"
        assert ws.cell(row=2, column=2).value == "OldMsg"
        # Red fill for REMOVED
        fill = ws.cell(row=2, column=1).fill
        assert fill.start_color.rgb == "00FFEEF0"
        wb.close()
