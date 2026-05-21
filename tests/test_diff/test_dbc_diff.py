"""Tests for core.diff.dbc_diff — DBC version diff engine."""

from core.diff.dbc_diff import DBCDiffEngine, DiffType
from core.parsers.dbc_parser import DBCData, MessageDef, SignalDef


def _make_signal(
    name,
    start_bit=0,
    bit_length=8,
    factor=1.0,
    offset=0.0,
    byte_order="little_endian",
    value_type="unsigned",
    unit="",
    minimum=0.0,
    maximum=255.0,
    receivers=None,
    comment="",
):
    return SignalDef(
        name=name,
        start_bit=start_bit,
        bit_length=bit_length,
        byte_order=byte_order,
        value_type=value_type,
        factor=factor,
        offset=offset,
        minimum=minimum,
        maximum=maximum,
        unit=unit,
        comment=comment,
        receivers=receivers or [],
        value_descriptions={},
        mux=None,
    )


def _make_message(name, msg_id=0x100, dlc=8, sender="VCU", signals=None):
    return MessageDef(
        id=msg_id,
        name=name,
        dlc=dlc,
        sender=sender,
        signals=signals or [],
        comment="",
        is_extended=False,
    )


def _make_dbc(messages, version="v1"):
    return DBCData(
        version=version,
        messages=messages,
        nodes=[],
        value_tables={},
        comments={},
        attributes={},
        source_path="<test>",
    )


# ── Core diff tests ──────────────────────────────────────────────────────────


class TestDBCDiffEngine:
    def setup_method(self):
        self.engine = DBCDiffEngine()

    def test_identical_dbc_no_diff(self):
        msg = _make_message("Msg1", signals=[_make_signal("Sig1")])
        old = _make_dbc([msg])
        new = _make_dbc([msg])
        result = self.engine.compare(old, new)
        assert result.summary["messages_added"] == 0
        assert result.summary["messages_removed"] == 0
        assert result.summary["messages_modified"] == 0

    def test_added_message(self):
        old = _make_dbc([])
        new_msg = _make_message("NewMsg", signals=[_make_signal("S1")])
        new = _make_dbc([new_msg])
        result = self.engine.compare(old, new)
        assert result.summary["messages_added"] == 1
        assert "NewMsg" in result.added_messages

    def test_removed_message(self):
        old_msg = _make_message("OldMsg")
        old = _make_dbc([old_msg])
        new = _make_dbc([])
        result = self.engine.compare(old, new)
        assert result.summary["messages_removed"] == 1
        assert "OldMsg" in result.removed_messages

    def test_added_signal(self):
        msg_old = _make_message("M1", signals=[_make_signal("S1")])
        msg_new = _make_message("M1", signals=[_make_signal("S1"), _make_signal("S2")])
        result = self.engine.compare(_make_dbc([msg_old]), _make_dbc([msg_new]))
        assert result.summary["signals_added"] == 1

    def test_removed_signal(self):
        msg_old = _make_message("M1", signals=[_make_signal("S1"), _make_signal("S2")])
        msg_new = _make_message("M1", signals=[_make_signal("S1")])
        result = self.engine.compare(_make_dbc([msg_old]), _make_dbc([msg_new]))
        assert result.summary["signals_removed"] == 1

    def test_modified_signal(self):
        sig_old = _make_signal("S1", factor=1.0)
        sig_new = _make_signal("S1", factor=0.5)
        result = self.engine.compare(
            _make_dbc([_make_message("M1", signals=[sig_old])]),
            _make_dbc([_make_message("M1", signals=[sig_new])]),
        )
        assert result.summary["signals_modified"] == 1
        md = result.message_diffs[0]
        sd = md.signal_diffs[0]
        assert sd.diff_type == DiffType.MODIFIED
        assert "factor" in sd.changes

    def test_modified_message_dlc(self):
        m1 = _make_message("M1", dlc=8)
        m2 = _make_message("M1", dlc=16)
        result = self.engine.compare(_make_dbc([m1]), _make_dbc([m2]))
        assert result.summary["messages_modified"] == 1
        assert "dlc" in result.message_diffs[0].changes

    def test_modified_sender(self):
        m1 = _make_message("M1", sender="VCU")
        m2 = _make_message("M1", sender="BMS")
        result = self.engine.compare(_make_dbc([m1]), _make_dbc([m2]))
        assert "sender" in result.message_diffs[0].changes

    def test_modified_comment(self):
        m1 = _make_message("M1")
        m1.comment = "old"
        m2 = _make_message("M1")
        m2.comment = "new"
        result = self.engine.compare(_make_dbc([m1]), _make_dbc([m2]))
        assert "comment" in result.message_diffs[0].changes

    def test_receivers_change(self):
        sig_old = _make_signal("S1", receivers=["A"])
        sig_new = _make_signal("S1", receivers=["B"])
        result = self.engine.compare(
            _make_dbc([_make_message("M1", signals=[sig_old])]),
            _make_dbc([_make_message("M1", signals=[sig_new])]),
        )
        sd = result.message_diffs[0].signal_diffs[0]
        assert "receivers" in sd.changes

    def test_value_descriptions_change(self):
        sig_old = _make_signal("S1")
        sig_old.value_descriptions = {0: "Off"}
        sig_new = _make_signal("S1")
        sig_new.value_descriptions = {0: "On"}
        result = self.engine.compare(
            _make_dbc([_make_message("M1", signals=[sig_old])]),
            _make_dbc([_make_message("M1", signals=[sig_new])]),
        )
        sd = result.message_diffs[0].signal_diffs[0]
        assert "value_descriptions" in sd.changes

    def test_text_report(self):
        old = _make_dbc([_make_message("M1", signals=[_make_signal("S1")])])
        new = _make_dbc([])
        result = self.engine.compare(old, new)
        report = self.engine.generate_text_report(result)
        assert "DBC Diff Report" in report
        assert "M1" in report

    def test_excel_report(self, tmp_path):
        old = _make_dbc([_make_message("M1", signals=[_make_signal("S1")])])
        new = _make_dbc([_make_message("M1", signals=[_make_signal("S1", factor=0.5)])])
        result = self.engine.compare(old, new)
        out = tmp_path / "diff.xlsx"
        self.engine.export_excel_report(result, out)
        assert out.exists()
        assert out.stat().st_size > 0

    def test_mixed_changes(self):
        """Added, removed, modified messages in one diff."""
        m1 = _make_message("Keep", signals=[_make_signal("S1")])
        m2 = _make_message("Remove")
        m3 = _make_message("Keep", signals=[_make_signal("S1", factor=2.0)])
        m4 = _make_message("Add")
        old = _make_dbc([m1, m2])
        new = _make_dbc([m3, m4])
        result = self.engine.compare(old, new)
        assert "Remove" in result.removed_messages
        assert "Add" in result.added_messages


# ── P1: Text report content verification ────────────────────────────────────


class TestDBCDiffTextReport:
    def setup_method(self):
        self.engine = DBCDiffEngine()

    def test_report_header_and_versions(self):
        old = _make_dbc([], version="v1.0")
        new = _make_dbc([], version="v2.0")
        result = self.engine.compare(old, new)
        report = self.engine.generate_text_report(result)
        assert "Old: v1.0" in report
        assert "New: v2.0" in report

    def test_report_summary_counts(self):
        sig = _make_signal("S1")
        old = _make_dbc([_make_message("M1", signals=[sig])])
        new = _make_dbc([_make_message("M1", signals=[_make_signal("S1", factor=0.5)])])
        result = self.engine.compare(old, new)
        report = self.engine.generate_text_report(result)
        assert "Signals modified:  1" in report

    def test_report_diff_markers(self):
        sig = _make_signal("S1")
        old = _make_dbc(
            [
                _make_message("Keep", signals=[sig]),
                _make_message("Removed"),
            ]
        )
        new = _make_dbc(
            [
                _make_message("Keep", signals=[_make_signal("S1", factor=2.0)]),
                _make_message("Added"),
            ]
        )
        result = self.engine.compare(old, new)
        report = self.engine.generate_text_report(result)
        assert "~" in report  # modified marker
        assert "+" in report  # added marker
        assert "-" in report  # removed marker

    def test_report_signal_field_change_line(self):
        old = _make_dbc([_make_message("M1", signals=[_make_signal("S1", factor=1.0)])])
        new = _make_dbc([_make_message("M1", signals=[_make_signal("S1", factor=0.5)])])
        result = self.engine.compare(old, new)
        report = self.engine.generate_text_report(result)
        assert "factor" in report
        assert "1.0" in report
        assert "0.5" in report
        assert "→" in report


# ── P1: Signal field change coverage ────────────────────────────────────────


class TestDBCDiffSignalFields:
    def setup_method(self):
        self.engine = DBCDiffEngine()

    def _diff_signal(self, old_sig, new_sig):
        old = _make_dbc([_make_message("M1", signals=[old_sig])])
        new = _make_dbc([_make_message("M1", signals=[new_sig])])
        result = self.engine.compare(old, new)
        return result.message_diffs[0].signal_diffs[0]

    def test_byte_order_change(self):
        old = _make_signal("S1", byte_order="little_endian")
        new = _make_signal("S1", byte_order="big_endian")
        sd = self._diff_signal(old, new)
        assert "byte_order" in sd.changes
        assert sd.changes["byte_order"] == ("little_endian", "big_endian")

    def test_bit_length_change(self):
        old = _make_signal("S1", bit_length=8)
        new = _make_signal("S1", bit_length=16)
        sd = self._diff_signal(old, new)
        assert "bit_length" in sd.changes
        assert sd.changes["bit_length"] == (8, 16)

    def test_offset_change(self):
        old = _make_signal("S1", offset=0.0)
        new = _make_signal("S1", offset=-500.0)
        sd = self._diff_signal(old, new)
        assert "offset" in sd.changes
        assert sd.changes["offset"] == (0.0, -500.0)

    def test_minimum_change(self):
        old = _make_signal("S1", minimum=0.0)
        new = _make_signal("S1", minimum=-100.0)
        sd = self._diff_signal(old, new)
        assert "minimum" in sd.changes

    def test_maximum_change(self):
        old = _make_signal("S1", maximum=255.0)
        new = _make_signal("S1", maximum=65535.0)
        sd = self._diff_signal(old, new)
        assert "maximum" in sd.changes

    def test_unit_change(self):
        old = _make_signal("S1", unit="rpm")
        new = _make_signal("S1", unit="rad/s")
        sd = self._diff_signal(old, new)
        assert "unit" in sd.changes
        assert sd.changes["unit"] == ("rpm", "rad/s")

    def test_value_type_change(self):
        old = _make_signal("S1", value_type="unsigned")
        new = _make_signal("S1", value_type="signed")
        sd = self._diff_signal(old, new)
        assert "value_type" in sd.changes

    def test_start_bit_change(self):
        old = _make_signal("S1", start_bit=0)
        new = _make_signal("S1", start_bit=8)
        sd = self._diff_signal(old, new)
        assert "start_bit" in sd.changes

    def test_no_change_same_signal(self):
        sig = _make_signal("S1", factor=0.1, offset=-40, unit="degC")
        old = _make_dbc([_make_message("M1", signals=[sig])])
        new = _make_dbc([_make_message("M1", signals=[sig])])
        result = self.engine.compare(old, new)
        assert result.summary["signals_modified"] == 0
        assert len(result.message_diffs) == 0


# ── P1: Excel report content verification ───────────────────────────────────


class TestDBCDiffExcelReport:
    def setup_method(self):
        self.engine = DBCDiffEngine()

    def test_excel_sheet_name_and_headers(self, tmp_path):
        from openpyxl import load_workbook

        old = _make_dbc([_make_message("M1", signals=[_make_signal("S1")])])
        new = _make_dbc([_make_message("M1", signals=[_make_signal("S1", factor=0.5)])])
        result = self.engine.compare(old, new)
        out = tmp_path / "diff.xlsx"
        self.engine.export_excel_report(result, out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.title == "Diff Report"
        headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
        assert headers == ["类型", "报文", "报文ID", "信号", "变更字段", "旧值", "新值"]
        wb.close()

    def test_excel_data_values(self, tmp_path):
        from openpyxl import load_workbook

        old = _make_dbc([_make_message("M1", signals=[_make_signal("S1", factor=1.0)])])
        new = _make_dbc([_make_message("M1", signals=[_make_signal("S1", factor=0.5)])])
        result = self.engine.compare(old, new)
        out = tmp_path / "diff.xlsx"
        self.engine.export_excel_report(result, out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "修改"
        assert ws.cell(row=2, column=2).value == "M1"
        assert ws.cell(row=2, column=4).value == "S1"
        assert ws.cell(row=2, column=5).value == "factor"
        wb.close()

    def test_excel_fill_colors(self, tmp_path):
        from openpyxl import load_workbook

        sig_old = _make_signal("S1")
        sig_new = _make_signal("S1", factor=0.5)
        old = _make_dbc(
            [
                _make_message("M1", signals=[sig_old]),
                _make_message("M2"),
                _make_message("M3", signals=[_make_signal("S2")]),
            ]
        )
        new = _make_dbc(
            [
                _make_message("M1", signals=[sig_new]),
                _make_message("M4"),
                _make_message("M3", signals=[_make_signal("S2")]),
            ]
        )
        result = self.engine.compare(old, new)
        out = tmp_path / "diff.xlsx"
        self.engine.export_excel_report(result, out)
        wb = load_workbook(out)
        ws = wb.active
        # Find rows by content
        for row in range(2, ws.max_row + 1):
            op = ws.cell(row=row, column=1).value
            fill = ws.cell(row=row, column=1).fill
            if op == "修改":
                assert fill.start_color.rgb == "00FFF5B1"  # yellow
            elif op == "新增":
                assert fill.start_color.rgb == "00E6FFED"  # green
            elif op == "删除":
                assert fill.start_color.rgb == "00FFEEF0"  # red
        wb.close()
