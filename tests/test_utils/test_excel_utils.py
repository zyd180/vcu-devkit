"""Tests for core.utils.excel_utils — shared Excel formatting."""

import pytest
from openpyxl import Workbook

from core.utils.excel_utils import (
    HEADER_FILL, HEADER_FONT, THIN_BORDER, CENTER_ALIGN,
    write_header_row, auto_width, apply_border,
)


class TestExcelUtils:

    def test_write_header_row(self):
        wb = Workbook()
        ws = wb.active
        write_header_row(ws, ["Col A", "Col B", "Col C"])
        assert ws.cell(1, 1).value == "Col A"
        assert ws.cell(1, 2).value == "Col B"
        assert ws.cell(1, 3).value == "Col C"
        # Check styling applied
        assert ws.cell(1, 1).fill == HEADER_FILL
        assert ws.cell(1, 1).font == HEADER_FONT

    def test_write_header_row_custom_row(self):
        wb = Workbook()
        ws = wb.active
        write_header_row(ws, ["A", "B"], row=5)
        assert ws.cell(5, 1).value == "A"
        assert ws.cell(5, 2).value == "B"

    def test_auto_width(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(1, 1).value = "Short"
        ws.cell(2, 1). value = "A much longer value"
        auto_width(ws)
        width = ws.column_dimensions["A"].width
        assert width >= len("A much longer value") + 3

    def test_auto_width_max_cap(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(1, 1).value = "x" * 100
        auto_width(ws, max_width=20)
        assert ws.column_dimensions["A"].width <= 20

    def test_apply_border(self):
        wb = Workbook()
        ws = wb.active
        for r in range(1, 4):
            for c in range(1, 3):
                ws.cell(r, c).value = f"r{r}c{c}"
        apply_border(ws, 1, 3, 2)
        for r in range(1, 4):
            for c in range(1, 3):
                assert ws.cell(r, c).border == THIN_BORDER

    def test_constants_not_none(self):
        assert HEADER_FILL is not None
        assert HEADER_FONT is not None
        assert THIN_BORDER is not None
        assert CENTER_ALIGN is not None
