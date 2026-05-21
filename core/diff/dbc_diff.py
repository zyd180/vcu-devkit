"""DBC version diff engine — signal-level comparison."""

from __future__ import annotations

from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Any

from core.parsers.dbc_parser import DBCData, MessageDef, SignalDef


class DiffType(Enum):
    ADDED = "added"
    REMOVED = "removed"
    MODIFIED = "modified"
    UNCHANGED = "unchanged"


@dataclass
class SignalDiff:
    """Diff result for a single signal."""

    signal_name: str
    diff_type: DiffType
    changes: dict[str, tuple[Any, Any]] = field(default_factory=dict)
    message_name: str = ""


@dataclass
class MessageDiff:
    """Diff result for a single message."""

    message_name: str
    diff_type: DiffType
    id: int = 0
    signal_diffs: list[SignalDiff] = field(default_factory=list)
    changes: dict[str, tuple[Any, Any]] = field(default_factory=dict)


@dataclass
class DBCDiffResult:
    """Complete DBC diff result."""

    old_version: str
    new_version: str
    message_diffs: list[MessageDiff] = field(default_factory=list)
    added_messages: list[str] = field(default_factory=list)
    removed_messages: list[str] = field(default_factory=list)
    summary: dict[str, int] = field(default_factory=dict)


# ── Fields to compare on signals ─────────────────────────────────────────────

SIGNAL_COMPARE_FIELDS = [
    "start_bit",
    "bit_length",
    "byte_order",
    "value_type",
    "factor",
    "offset",
    "minimum",
    "maximum",
    "unit",
    "comment",
]


class DBCDiffEngine:
    """Compare two DBCData instances at signal granularity."""

    def compare(self, old: DBCData, new: DBCData) -> DBCDiffResult:
        old_msgs = {m.name: m for m in old.messages}
        new_msgs = {m.name: m for m in new.messages}

        all_names = set(old_msgs.keys()) | set(new_msgs.keys())
        message_diffs: list[MessageDiff] = []
        added_messages: list[str] = []
        removed_messages: list[str] = []

        for name in sorted(all_names):
            old_msg = old_msgs.get(name)
            new_msg = new_msgs.get(name)

            if old_msg is None and new_msg is not None:
                # Entire message added
                added_messages.append(name)
                sdiffs = [
                    SignalDiff(
                        signal_name=s.name,
                        diff_type=DiffType.ADDED,
                        message_name=name,
                    )
                    for s in new_msg.signals
                ]
                message_diffs.append(
                    MessageDiff(
                        message_name=name,
                        diff_type=DiffType.ADDED,
                        id=new_msg.id,
                        signal_diffs=sdiffs,
                    )
                )
            elif old_msg is not None and new_msg is None:
                # Entire message removed
                removed_messages.append(name)
                sdiffs = [
                    SignalDiff(
                        signal_name=s.name,
                        diff_type=DiffType.REMOVED,
                        message_name=name,
                    )
                    for s in old_msg.signals
                ]
                message_diffs.append(
                    MessageDiff(
                        message_name=name,
                        diff_type=DiffType.REMOVED,
                        id=old_msg.id,
                        signal_diffs=sdiffs,
                    )
                )
            else:
                # Both exist — compare signals
                sd = self._compare_signals(old_msg, new_msg)
                msg_changes = self._compare_message_fields(old_msg, new_msg)
                diff_type = DiffType.MODIFIED if sd or msg_changes else DiffType.UNCHANGED
                if diff_type != DiffType.UNCHANGED:
                    message_diffs.append(
                        MessageDiff(
                            message_name=name,
                            diff_type=diff_type,
                            id=new_msg.id,
                            signal_diffs=sd,
                            changes=msg_changes,
                        )
                    )

        added_count = sum(1 for md in message_diffs if md.diff_type == DiffType.ADDED)
        removed_count = sum(1 for md in message_diffs if md.diff_type == DiffType.REMOVED)
        modified_count = sum(1 for md in message_diffs if md.diff_type == DiffType.MODIFIED)
        signals_added = sum(1 for md in message_diffs for sd in md.signal_diffs if sd.diff_type == DiffType.ADDED)
        signals_removed = sum(1 for md in message_diffs for sd in md.signal_diffs if sd.diff_type == DiffType.REMOVED)
        signals_modified = sum(1 for md in message_diffs for sd in md.signal_diffs if sd.diff_type == DiffType.MODIFIED)

        return DBCDiffResult(
            old_version=old.version or old.source_path,
            new_version=new.version or new.source_path,
            message_diffs=message_diffs,
            added_messages=added_messages,
            removed_messages=removed_messages,
            summary={
                "messages_added": added_count,
                "messages_removed": removed_count,
                "messages_modified": modified_count,
                "signals_added": signals_added,
                "signals_removed": signals_removed,
                "signals_modified": signals_modified,
            },
        )

    def _compare_signals(self, old_msg: MessageDef, new_msg: MessageDef) -> list[SignalDiff]:
        old_sigs = {s.name: s for s in old_msg.signals}
        new_sigs = {s.name: s for s in new_msg.signals}
        all_names = set(old_sigs.keys()) | set(new_sigs.keys())
        result: list[SignalDiff] = []

        for name in sorted(all_names):
            old_sig = old_sigs.get(name)
            new_sig = new_sigs.get(name)

            if old_sig is None and new_sig is not None:
                result.append(
                    SignalDiff(
                        signal_name=name,
                        diff_type=DiffType.ADDED,
                        message_name=new_msg.name,
                    )
                )
            elif old_sig is not None and new_sig is None:
                result.append(
                    SignalDiff(
                        signal_name=name,
                        diff_type=DiffType.REMOVED,
                        message_name=old_msg.name,
                    )
                )
            else:
                changes = self._compare_signal_fields(old_sig, new_sig)
                if changes:
                    result.append(
                        SignalDiff(
                            signal_name=name,
                            diff_type=DiffType.MODIFIED,
                            changes=changes,
                            message_name=new_msg.name,
                        )
                    )

        return result

    @staticmethod
    def _compare_signal_fields(old: SignalDef, new: SignalDef) -> dict[str, tuple[Any, Any]]:
        changes: dict[str, tuple[Any, Any]] = {}
        for field_name in SIGNAL_COMPARE_FIELDS:
            old_val = getattr(old, field_name)
            new_val = getattr(new, field_name)
            if old_val != new_val:
                changes[field_name] = (old_val, new_val)
        # Compare receivers as sets
        if set(old.receivers) != set(new.receivers):
            changes["receivers"] = (old.receivers, new.receivers)
        # Compare value descriptions
        if old.value_descriptions != new.value_descriptions:
            changes["value_descriptions"] = (old.value_descriptions, new.value_descriptions)
        return changes

    @staticmethod
    def _compare_message_fields(old: MessageDef, new: MessageDef) -> dict[str, tuple[Any, Any]]:
        changes: dict[str, tuple[Any, Any]] = {}
        if old.dlc != new.dlc:
            changes["dlc"] = (old.dlc, new.dlc)
        if old.sender != new.sender:
            changes["sender"] = (old.sender, new.sender)
        if old.comment != new.comment:
            changes["comment"] = (old.comment, new.comment)
        return changes

    def generate_text_report(self, diff: DBCDiffResult) -> str:
        """Generate a human-readable text diff report."""
        lines = [
            "DBC Diff Report",
            f"Old: {diff.old_version}",
            f"New: {diff.new_version}",
            "",
            "Summary:",
            f"  Messages added:    {diff.summary.get('messages_added', 0)}",
            f"  Messages removed:  {diff.summary.get('messages_removed', 0)}",
            f"  Messages modified: {diff.summary.get('messages_modified', 0)}",
            f"  Signals added:     {diff.summary.get('signals_added', 0)}",
            f"  Signals removed:   {diff.summary.get('signals_removed', 0)}",
            f"  Signals modified:  {diff.summary.get('signals_modified', 0)}",
            "",
        ]

        for md in diff.message_diffs:
            prefix = {"added": "+", "removed": "-", "modified": "~"}.get(md.diff_type.value, " ")
            lines.append(f"{prefix} Message: {md.message_name} (0x{md.id:03X})")
            for sd in md.signal_diffs:
                sp = {"added": "  +", "removed": "  -", "modified": "  ~"}.get(sd.diff_type.value, "   ")
                lines.append(f"{sp} {sd.signal_name}")
                for field_name, (old, new) in sd.changes.items():
                    lines.append(f"      {field_name}: {old} → {new}")

        return "\n".join(lines)

    def export_excel_report(self, diff: DBCDiffResult, output_path: Path):
        """Export diff report to Excel."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Diff Report"

        green_fill = PatternFill(start_color="E6FFED", fill_type="solid")
        red_fill = PatternFill(start_color="FFEEF0", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF5B1", fill_type="solid")

        headers = ["类型", "报文", "报文ID", "信号", "变更字段", "旧值", "新值"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)

        row = 2
        for md in diff.message_diffs:
            if md.signal_diffs:
                for sd in md.signal_diffs:
                    prefix = {"added": "新增", "removed": "删除", "modified": "修改"}.get(sd.diff_type.value, "")
                    ws.cell(row=row, column=1, value=prefix)
                    ws.cell(row=row, column=2, value=md.message_name)
                    ws.cell(row=row, column=3, value=f"0x{md.id:03X}")
                    ws.cell(row=row, column=4, value=sd.signal_name)
                    if sd.changes:
                        for field_name, (old, new) in sd.changes.items():
                            ws.cell(row=row, column=5, value=field_name)
                            ws.cell(row=row, column=6, value=str(old))
                            ws.cell(row=row, column=7, value=str(new))
                            fill = {"added": green_fill, "removed": red_fill, "modified": yellow_fill}.get(
                                sd.diff_type.value
                            )
                            if fill:
                                for c in range(1, 8):
                                    ws.cell(row=row, column=c).fill = fill
                            row += 1
                    else:
                        fill = {"added": green_fill, "removed": red_fill}.get(sd.diff_type.value)
                        if fill:
                            for c in range(1, 8):
                                ws.cell(row=row, column=c).fill = fill
                        row += 1
            else:
                prefix = {"added": "新增报文", "removed": "删除报文"}.get(md.diff_type.value, "修改")
                ws.cell(row=row, column=1, value=prefix)
                ws.cell(row=row, column=2, value=md.message_name)
                ws.cell(row=row, column=3, value=f"0x{md.id:03X}")
                fill = {"added": green_fill, "removed": red_fill}.get(md.diff_type.value)
                if fill:
                    for c in range(1, 8):
                        ws.cell(row=row, column=c).fill = fill
                row += 1

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        wb.save(str(output_path))
