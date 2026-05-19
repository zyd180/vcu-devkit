"""Excel report generator for DBC data and diff results."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from core.parsers.dbc_parser import DBCData
from core.diff.dbc_diff import DBCDiffResult, DiffType


class ReportGenerator:
    """Generate Excel reports from DBC data."""

    def generate_signal_matrix(self, data: DBCData, output_path: Path):
        """Export complete signal matrix to Excel."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment
        from core.utils.excel_utils import write_header_row, auto_width, THIN_BORDER

        wb = Workbook()
        ws = wb.active
        ws.title = "Signal Matrix"

        headers = [
            "报文名", "CAN ID", "DLC", "发送方", "信号名", "起始位", "位长",
            "字节序", "值类型", "Factor", "Offset", "最小值", "最大值", "单位", "接收方", "值描述",
        ]
        write_header_row(ws, headers)

        row = 2
        for msg in data.messages:
            for sig in msg.signals:
                vals = [
                    msg.name,
                    f"0x{msg.id:03X}",
                    msg.dlc,
                    msg.sender,
                    sig.name,
                    sig.start_bit,
                    sig.bit_length,
                    sig.byte_order,
                    sig.value_type,
                    sig.factor,
                    sig.offset,
                    sig.minimum,
                    sig.maximum,
                    sig.unit,
                    ", ".join(sig.receivers),
                    "; ".join(f"{v}={d}" for v, d in sorted(sig.value_descriptions.items())),
                ]
                for col, val in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = THIN_BORDER
                row += 1

        auto_width(ws)

        # Freeze header
        ws.freeze_panes = "A2"

        wb.save(str(output_path))

    def generate_message_summary(self, data: DBCData, output_path: Path):
        """Export message-level summary."""
        from openpyxl import Workbook
        from core.utils.excel_utils import write_header_row, auto_width, THIN_BORDER

        wb = Workbook()
        ws = wb.active
        ws.title = "Message Summary"

        headers = ["报文名", "CAN ID", "DLC (bytes)", "发送方", "信号数量", "备注"]
        write_header_row(ws, headers)

        for row, msg in enumerate(data.messages, 2):
            ws.cell(row=row, column=1, value=msg.name)
            ws.cell(row=row, column=2, value=f"0x{msg.id:03X}")
            ws.cell(row=row, column=3, value=msg.dlc)
            ws.cell(row=row, column=4, value=msg.sender)
            ws.cell(row=row, column=5, value=len(msg.signals))
            ws.cell(row=row, column=6, value=msg.comment)

        auto_width(ws)
        ws.freeze_panes = "A2"
        wb.save(str(output_path))

    def generate_diff_report(self, diff: DBCDiffResult, output_path: Path):
        """Export diff report to Excel with formatting."""
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font
        from core.utils.excel_utils import write_header_row, auto_width

        wb = Workbook()

        # ── Sheet 1: Summary ──
        ws_sum = wb.active
        ws_sum.title = "变更汇总"
        ws_sum.cell(row=1, column=1, value="DBC版本对比报告").font = Font(bold=True, size=14)
        ws_sum.cell(row=2, column=1, value=f"旧版本: {diff.old_version}")
        ws_sum.cell(row=3, column=1, value=f"新版本: {diff.new_version}")
        ws_sum.cell(row=4, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        row = 6
        for k, v in diff.summary.items():
            label = {
                "messages_added": "新增报文",
                "messages_removed": "删除报文",
                "messages_modified": "修改报文",
                "signals_added": "新增信号",
                "signals_removed": "删除信号",
                "signals_modified": "修改信号",
            }.get(k, k)
            ws_sum.cell(row=row, column=1, value=label)
            ws_sum.cell(row=row, column=2, value=v)
            row += 1

        # ── Sheet 2: Detail ──
        ws_det = wb.create_sheet("变更详情")
        green = PatternFill(start_color="E6FFED", fill_type="solid")
        red = PatternFill(start_color="FFEEF0", fill_type="solid")
        yellow = PatternFill(start_color="FFF5B1", fill_type="solid")
        bold = Font(bold=True)

        headers = ["操作", "报文", "CAN ID", "信号", "变更字段", "旧值", "新值"]
        write_header_row(ws_det, headers)

        row = 2
        fill_map = {
            DiffType.ADDED: green,
            DiffType.REMOVED: red,
            DiffType.MODIFIED: yellow,
        }
        op_map = {
            DiffType.ADDED: "新增",
            DiffType.REMOVED: "删除",
            DiffType.MODIFIED: "修改",
        }

        for md in diff.message_diffs:
            if md.signal_diffs:
                for sd in md.signal_diffs:
                    fill = fill_map.get(sd.diff_type)
                    if sd.changes:
                        for field_name, (old_val, new_val) in sd.changes.items():
                            ws_det.cell(row=row, column=1, value=op_map.get(sd.diff_type, ""))
                            ws_det.cell(row=row, column=2, value=md.message_name)
                            ws_det.cell(row=row, column=3, value=f"0x{md.id:03X}")
                            ws_det.cell(row=row, column=4, value=sd.signal_name)
                            ws_det.cell(row=row, column=5, value=field_name)
                            ws_det.cell(row=row, column=6, value=str(old_val))
                            ws_det.cell(row=row, column=7, value=str(new_val))
                            if fill:
                                for c in range(1, 8):
                                    ws_det.cell(row=row, column=c).fill = fill
                            row += 1
                    else:
                        ws_det.cell(row=row, column=1, value=op_map.get(sd.diff_type, ""))
                        ws_det.cell(row=row, column=2, value=md.message_name)
                        ws_det.cell(row=row, column=3, value=f"0x{md.id:03X}")
                        ws_det.cell(row=row, column=4, value=sd.signal_name)
                        if fill:
                            for c in range(1, 8):
                                ws_det.cell(row=row, column=c).fill = fill
                        row += 1
            else:
                fill = fill_map.get(md.diff_type)
                ws_det.cell(row=row, column=1, value=op_map.get(md.diff_type, ""))
                ws_det.cell(row=row, column=2, value=md.message_name)
                ws_det.cell(row=row, column=3, value=f"0x{md.id:03X}")
                if fill:
                    for c in range(1, 8):
                        ws_det.cell(row=row, column=c).fill = fill
                row += 1

        auto_width(ws_det)

        ws_det.freeze_panes = "A2"
        wb.save(str(output_path))
