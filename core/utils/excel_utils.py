"""Shared Excel formatting utilities for openpyxl."""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill(start_color="1A73E8", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
CENTER_ALIGN = Alignment(horizontal="center")


def write_header_row(ws: Worksheet, headers: list[str], row: int = 1) -> None:
    """Write a styled header row to a worksheet."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def auto_width(ws: Worksheet, max_width: int = 40) -> None:
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col_cells), default=8)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, max_width)


def apply_border(ws: Worksheet, start_row: int, end_row: int, col_count: int) -> None:
    """Apply thin border to a range of cells."""
    for row in range(start_row, end_row + 1):
        for col in range(1, col_count + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
