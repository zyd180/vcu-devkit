"""Traceability Matrix business logic controller."""

from __future__ import annotations

import json
import logging
from pathlib import Path

from core.db.crud_mixin import CRUDMixin
from core.db.manager import DatabaseManager
from core.db.models import Requirement, TraceabilityLink

logger = logging.getLogger(__name__)


class TraceMatrixController(CRUDMixin):
    """Manage requirement ↔ artifact traceability links."""

    model = Requirement
    LINK_TYPES = ["swc", "signal", "dtc", "test_case", "port", "runnable"]

    def __init__(self, db_manager: DatabaseManager | None = None):
        self.db = db_manager or DatabaseManager()
        if not self.db.is_ready():
            self.db.init()

    # ── Requirement CRUD ───────────────────────────────────────────────────

    def get_requirements(self, module: str | None = None) -> list[Requirement]:
        return self.db.get_requirements(module=module)

    def get_req_by_id(self, req_id: str) -> Requirement | None:
        try:
            return Requirement.get(Requirement.req_id == req_id)
        except Requirement.DoesNotExist:
            return None

    def add_requirement(self, req_id: str, title: str, **kwargs) -> Requirement | None:
        try:
            return self.db.add_requirement(req_id=req_id, title=title, **kwargs)
        except Exception as exc:
            logger.error("Failed to add requirement %s: %s", req_id, exc)
            return None

    def update_requirement(self, db_id: int, **kwargs) -> bool:
        return self.update_fields(db_id, **kwargs)

    def remove_requirement(self, db_id: int) -> bool:
        try:
            req = Requirement.get_by_id(db_id)
            # Remove all associated links
            TraceabilityLink.delete().where(TraceabilityLink.req == req).execute()
            req.delete_instance()
            return True
        except Requirement.DoesNotExist:
            return False

    def get_requirement_as_dict(self, db_id: int) -> dict | None:
        try:
            req = Requirement.get_by_id(db_id)
            links = list(TraceabilityLink.select().where(TraceabilityLink.req == req))
            return {
                "id": req.id,
                "req_id": req.req_id,
                "title": req.title,
                "description": req.description or "",
                "source": req.source,
                "module_name": req.module_name or "",
                "status": req.status,
                "links": [
                    {
                        "type": lnk.link_type,
                        "target": lnk.link_target,
                        "target_id": lnk.link_target_id or "",
                        "auto_matched": lnk.auto_matched,
                        "verified": lnk.verified,
                    }
                    for lnk in links
                ],
            }
        except Requirement.DoesNotExist:
            return None

    # ── Bulk import ────────────────────────────────────────────────────────

    def import_requirements(self, req_list: list[dict]) -> tuple[int, int]:
        """Import requirements from list of dicts. Returns (imported, skipped)."""
        imported = 0
        skipped = 0
        for item in req_list:
            req_id = item.get("req_id", "")
            if not req_id:
                skipped += 1
                continue
            if self.get_req_by_id(req_id):
                skipped += 1
                continue
            result = self.add_requirement(
                req_id=req_id,
                title=item.get("title", ""),
                description=item.get("description", ""),
                source=item.get("source", "import"),
                module_name=item.get("module_name"),
            )
            if result:
                imported += 1
            else:
                skipped += 1
        return imported, skipped

    # ── Link management ────────────────────────────────────────────────────

    def add_link(
        self, req_id: str, link_type: str, link_target: str, link_target_id: str = "", auto: bool = False
    ) -> bool:
        """Add a traceability link from requirement to artifact."""
        req = self.get_req_by_id(req_id)
        if req is None:
            return False
        # Check duplicate
        existing = (
            TraceabilityLink.select()
            .where(
                (TraceabilityLink.req == req)
                & (TraceabilityLink.link_type == link_type)
                & (TraceabilityLink.link_target == link_target)
            )
            .first()
        )
        if existing:
            return False
        TraceabilityLink.create(
            req=req,
            link_type=link_type,
            link_target=link_target,
            link_target_id=link_target_id,
            auto_matched=auto,
        )
        return True

    def remove_link(self, link_id: int) -> bool:
        try:
            link = TraceabilityLink.get_by_id(link_id)
            link.delete_instance()
            return True
        except TraceabilityLink.DoesNotExist:
            return False

    def verify_link(self, link_id: int) -> bool:
        try:
            link = TraceabilityLink.get_by_id(link_id)
            link.verified = True
            link.save()
            return True
        except TraceabilityLink.DoesNotExist:
            return False

    def get_links_for_req(self, req_id: str) -> list[dict]:
        req = self.get_req_by_id(req_id)
        if req is None:
            return []
        links = list(TraceabilityLink.select().where(TraceabilityLink.req == req))
        return [
            {
                "id": lnk.id,
                "type": lnk.link_type,
                "target": lnk.link_target,
                "target_id": lnk.link_target_id or "",
                "auto_matched": lnk.auto_matched,
                "verified": lnk.verified,
            }
            for lnk in links
        ]

    # ── Auto-matching ──────────────────────────────────────────────────────

    def auto_match_by_naming(self, artifacts: dict[str, list[str]]) -> int:
        """Auto-match requirements to artifacts by keyword similarity.

        artifacts: {"swc": ["SwcPowerMgr", ...], "signal": [...], "dtc": [...], "test_case": [...]}

        Returns number of new links created.
        """
        count = 0
        for req in self.get_requirements():
            req_keywords = self._extract_keywords(req.req_id + " " + req.title)
            for link_type, targets in artifacts.items():
                for target in targets:
                    target_keywords = self._extract_keywords(target)
                    if req_keywords & target_keywords:
                        # Check if link already exists
                        existing = (
                            TraceabilityLink.select()
                            .where(
                                (TraceabilityLink.req == req)
                                & (TraceabilityLink.link_type == link_type)
                                & (TraceabilityLink.link_target == target)
                            )
                            .first()
                        )
                        if not existing:
                            TraceabilityLink.create(
                                req=req,
                                link_type=link_type,
                                link_target=target,
                                auto_matched=True,
                            )
                            count += 1
        return count

    @staticmethod
    def _extract_keywords(text: str) -> set[str]:
        """Extract meaningful keywords from a name/title."""
        import re

        # Split on common separators, keep tokens >= 3 chars
        tokens = re.split(r"[_\s\-/.,;:()[\]{}]+", text)
        keywords = set()
        for t in tokens:
            t = t.strip().lower()
            if len(t) >= 3:
                keywords.add(t)
            # Also add camelCase segments
            segments = re.findall(r"[A-Z][a-z]+|[a-z]+|[A-Z]+(?=[A-Z]|$)", t)
            for seg in segments:
                seg = seg.lower()
                if len(seg) >= 3:
                    keywords.add(seg)
        return keywords

    # ── Matrix generation ──────────────────────────────────────────────────

    def get_matrix(self) -> dict:
        """Generate full traceability matrix (single query with LEFT JOIN).

        Returns: {req_id: {"title": str, "module": str, "links": {type: [targets]}}}
        """
        from peewee import JOIN

        # Single query: all requirements LEFT JOIN their links
        rows = (
            Requirement.select(Requirement, TraceabilityLink)
            .join(TraceabilityLink, JOIN.LEFT_OUTER, on=(TraceabilityLink.req == Requirement.id))
            .dicts()
        )

        # Python-side grouping
        matrix: dict[str, dict] = {}
        for row in rows:
            req_id = row["req_id"]
            if req_id not in matrix:
                matrix[req_id] = {
                    "title": row["title"],
                    "module": row.get("module_name") or "",
                    "links": {},
                    "link_count": 0,
                    "verified_count": 0,
                }
            # row may have link fields if joined, or be NULL if no links
            if row.get("link_type"):
                entry = matrix[req_id]
                entry["links"].setdefault(row["link_type"], []).append(row["link_target"])
                entry["link_count"] += 1
                if row.get("verified"):
                    entry["verified_count"] += 1
        return matrix

    def get_statistics(self) -> dict:
        """Get traceability statistics (4 queries instead of 6)."""
        # Query 1-2: requirement counts
        total_reqs = Requirement.select().count()
        linked_reqs = (
            Requirement.select()
            .where(Requirement.id.in_(TraceabilityLink.select(TraceabilityLink.req).distinct()))
            .count()
        )

        # Query 3: link counts (total, verified, auto) — single pass
        links = list(
            TraceabilityLink.select(
                TraceabilityLink.link_type,
                TraceabilityLink.verified,
                TraceabilityLink.auto_matched,
            )
        )
        total_links = len(links)
        verified_links = sum(1 for lnk in links if lnk.verified)
        auto_links = sum(1 for lnk in links if lnk.auto_matched)

        # Type counts from the same data
        type_counts: dict[str, int] = {}
        for link in links:
            type_counts[link.link_type] = type_counts.get(link.link_type, 0) + 1

        return {
            "total_requirements": total_reqs,
            "linked_requirements": linked_reqs,
            "unlinked_requirements": total_reqs - linked_reqs,
            "coverage_pct": (linked_reqs / total_reqs * 100) if total_reqs > 0 else 0,
            "total_links": total_links,
            "verified_links": verified_links,
            "auto_links": auto_links,
            "manual_links": total_links - auto_links,
            "links_by_type": type_counts,
        }

    def get_unlinked_requirements(self) -> list[Requirement]:
        """Get requirements with no traceability links."""
        linked_ids = set(lnk.req_id for lnk in TraceabilityLink.select(TraceabilityLink.req).distinct())
        return [r for r in self.get_requirements() if r.id not in linked_ids]

    def get_gaps(self) -> list[dict]:
        """Identify traceability gaps (requirements without full coverage)."""
        gaps = []
        for req in self.get_requirements():
            links = self.get_links_for_req(req.req_id)
            link_types = {lnk["type"] for lnk in links}
            missing = []
            if "swc" not in link_types and "signal" not in link_types:
                missing.append("implementation (SWC/Signal)")
            if "test_case" not in link_types:
                missing.append("test case")
            if missing:
                gaps.append(
                    {
                        "req_id": req.req_id,
                        "title": req.title,
                        "missing": missing,
                        "existing_types": list(link_types),
                    }
                )
        return gaps

    # ── Import / Export ─────────────────────────────────────────────────────

    def export_json(self, output_path: Path) -> tuple[bool, list[str]]:
        try:
            data = {
                "requirements": [self.get_requirement_as_dict(r.id) for r in self.get_requirements()],
                "statistics": self.get_statistics(),
                "gaps": self.get_gaps(),
            }
            output_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
            return True, []
        except Exception as exc:
            return False, [str(exc)]

    def import_json(self, input_path: Path) -> tuple[bool, list[str]]:
        try:
            data = json.loads(input_path.read_text(encoding="utf-8"))
            imported, skipped = self.import_requirements(data.get("requirements", []))
            # Import links
            links_imported = 0
            for req_data in data.get("requirements", []):
                req_id = req_data.get("req_id", "")
                for link in req_data.get("links", []):
                    if self.add_link(
                        req_id,
                        link.get("type", ""),
                        link.get("target", ""),
                        link.get("target_id", ""),
                        auto=link.get("auto_matched", False),
                    ):
                        links_imported += 1
            return True, [f"Imported {imported} requirements, {links_imported} links, {skipped} skipped"]
        except Exception as exc:
            return False, [str(exc)]

    def export_excel(self, output_path: Path) -> tuple[bool, list[str]]:
        """Export traceability matrix to Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            from core.utils.excel_utils import THIN_BORDER, auto_width, write_header_row

            wb = Workbook()

            # Sheet 1: Matrix
            ws = wb.active
            ws.title = "追溯矩阵"

            green = PatternFill(start_color="E6FFED", fill_type="solid")
            red = PatternFill(start_color="FFEEF0", fill_type="solid")

            headers = ["需求ID", "需求标题", "模块", "SWC", "信号", "DTC", "测试用例", "链接总数", "已验证"]
            write_header_row(ws, headers)

            matrix = self.get_matrix()
            for row, (req_id, info) in enumerate(sorted(matrix.items()), 2):
                ws.cell(row=row, column=1, value=req_id).border = THIN_BORDER
                ws.cell(row=row, column=2, value=info["title"]).border = THIN_BORDER
                ws.cell(row=row, column=3, value=info["module"]).border = THIN_BORDER
                ws.cell(row=row, column=4, value=", ".join(info["links"].get("swc", []))).border = THIN_BORDER
                ws.cell(row=row, column=5, value=", ".join(info["links"].get("signal", []))).border = THIN_BORDER
                ws.cell(row=row, column=6, value=", ".join(info["links"].get("dtc", []))).border = THIN_BORDER
                ws.cell(row=row, column=7, value=", ".join(info["links"].get("test_case", []))).border = THIN_BORDER
                ws.cell(row=row, column=8, value=info["link_count"]).border = THIN_BORDER
                ws.cell(row=row, column=9, value=info["verified_count"]).border = THIN_BORDER

                # Color code: green if fully linked, red if not
                fill = green if info["link_count"] > 0 else red
                for col in range(1, 10):
                    ws.cell(row=row, column=col).fill = fill

            auto_width(ws)
            ws.freeze_panes = "A2"

            # Sheet 2: Statistics
            ws2 = wb.create_sheet("统计")
            stats = self.get_statistics()
            ws2.cell(row=1, column=1, value="追溯统计").font = Font(bold=True, size=14)
            ws2.cell(row=3, column=1, value="总需求数")
            ws2.cell(row=3, column=2, value=stats["total_requirements"])
            ws2.cell(row=4, column=1, value="已追溯需求")
            ws2.cell(row=4, column=2, value=stats["linked_requirements"])
            ws2.cell(row=5, column=1, value="未追溯需求")
            ws2.cell(row=5, column=2, value=stats["unlinked_requirements"])
            ws2.cell(row=6, column=1, value="追溯覆盖率")
            ws2.cell(row=6, column=2, value=f"{stats['coverage_pct']:.1f}%")
            ws2.cell(row=8, column=1, value="总链接数")
            ws2.cell(row=8, column=2, value=stats["total_links"])
            ws2.cell(row=9, column=1, value="自动匹配")
            ws2.cell(row=9, column=2, value=stats["auto_links"])
            ws2.cell(row=10, column=1, value="已验证")
            ws2.cell(row=10, column=2, value=stats["verified_links"])

            # Sheet 3: Gaps
            ws3 = wb.create_sheet("追溯缺口")
            ws3.cell(row=1, column=1, value="需求ID").font = Font(bold=True)
            ws3.cell(row=1, column=2, value="需求标题").font = Font(bold=True)
            ws3.cell(row=1, column=3, value="缺失链接").font = Font(bold=True)
            for i, gap in enumerate(self.get_gaps(), 2):
                ws3.cell(row=i, column=1, value=gap["req_id"])
                ws3.cell(row=i, column=2, value=gap["title"])
                ws3.cell(row=i, column=3, value=", ".join(gap["missing"]))

            wb.save(str(output_path))
            return True, []
        except (OSError, ValueError) as exc:
            return False, [str(exc)]
