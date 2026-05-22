"""Test Generator business logic controller."""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path

from core.parsers.dbc_parser import DBCData, DBCParser, MessageDef, SignalDef


class TestMethod(Enum):
    BOUNDARY_VALUE = "boundary_value"
    EQUIVALENCE_PARTITION = "equivalence_partition"
    NORMAL_RANGE = "normal_range"
    ERROR_INJECTION = "error_injection"
    SIGNAL_TIMEOUT = "signal_timeout"
    E2E_PROTECTION = "e2e_protection"
    COUNTER_VALIDATION = "counter_validation"


@dataclass
class TestCase:
    """Single test case definition."""

    id: str
    name: str
    category: str
    method: str
    description: str
    preconditions: list[str] = field(default_factory=list)
    steps: list[str] = field(default_factory=list)
    expected_results: list[str] = field(default_factory=list)
    signal_name: str = ""
    message_name: str = ""
    input_value: str = ""
    priority: str = "medium"  # low / medium / high / critical
    status: str = "draft"  # draft / ready / passed / failed / blocked


class TestGeneratorController:
    """Generate and manage test cases from DBC/ARXML/DTC data."""

    def __init__(self):
        self.dbc_parser = DBCParser()
        self.current_dbc: DBCData | None = None
        self.test_cases: list[TestCase] = []
        self._counter = 0

    # ── DBC loading ────────────────────────────────────────────────────────

    def load_dbc(self, file_path: Path) -> tuple[bool, list[str]]:
        result = self.dbc_parser.parse(file_path)
        if result.success:
            self.current_dbc = result.data
            return True, []
        return False, result.errors

    # ── Test case generation ───────────────────────────────────────────────

    def generate_signal_tests(self, methods: list[TestMethod] | None = None) -> int:
        """Generate test cases from all signals in the loaded DBC."""
        if self.current_dbc is None:
            return 0
        if methods is None:
            methods = [TestMethod.BOUNDARY_VALUE, TestMethod.NORMAL_RANGE]

        count = 0
        for msg in self.current_dbc.messages:
            for sig in msg.signals:
                for method in methods:
                    cases = self._generate_for_signal(msg, sig, method)
                    self.test_cases.extend(cases)
                    count += len(cases)
        return count

    def _generate_for_signal(self, msg: MessageDef, sig: SignalDef, method: TestMethod) -> list[TestCase]:
        cases = []
        prefix = f"TC_{msg.name}_{sig.name}"
        sig_min = sig.minimum if sig.minimum is not None else 0
        sig_max = sig.maximum if sig.maximum is not None else 0
        sig_factor = sig.factor if sig.factor is not None else 1
        sig_unit = sig.unit or ""

        if method == TestMethod.BOUNDARY_VALUE:
            # Min boundary
            cases.append(
                TestCase(
                    id=f"{prefix}_BV_MIN",
                    name=f"{sig.name} 最小边界值",
                    category="边界值测试",
                    method=method.value,
                    description=f"验证信号 {sig.name} 在最小物理值 ({sig_min}) 时的行为",
                    steps=[
                        f"设置 {msg.name}.{sig.name} = {sig_min} {sig_unit}",
                        f"发送报文 {msg.name} (0x{msg.id:03X})",
                        "观察接收方响应",
                    ],
                    expected_results=[
                        f"接收方正确解析 {sig.name} = {sig_min} {sig_unit}",
                        "无错误或异常行为",
                    ],
                    signal_name=sig.name,
                    message_name=msg.name,
                    input_value=str(sig_min),
                    priority="high",
                )
            )
            # Max boundary
            cases.append(
                TestCase(
                    id=f"{prefix}_BV_MAX",
                    name=f"{sig.name} 最大边界值",
                    category="边界值测试",
                    method=method.value,
                    description=f"验证信号 {sig.name} 在最大物理值 ({sig_max}) 时的行为",
                    steps=[
                        f"设置 {msg.name}.{sig.name} = {sig_max} {sig_unit}",
                        f"发送报文 {msg.name} (0x{msg.id:03X})",
                        "观察接收方响应",
                    ],
                    expected_results=[
                        f"接收方正确解析 {sig.name} = {sig_max} {sig_unit}",
                        "无错误或异常行为",
                    ],
                    signal_name=sig.name,
                    message_name=msg.name,
                    input_value=str(sig_max),
                    priority="high",
                )
            )
            # Just below min
            cases.append(
                TestCase(
                    id=f"{prefix}_BV_BELOW",
                    name=f"{sig.name} 低于最小值",
                    category="边界值测试",
                    method=method.value,
                    description=f"验证信号 {sig.name} 在低于最小物理值时的处理",
                    steps=[
                        f"设置 {msg.name}.{sig.name} = {sig_min - sig_factor} {sig_unit}",
                        f"发送报文 {msg.name} (0x{msg.id:03X})",
                    ],
                    expected_results=[
                        "系统正确处理下溢值（截断或报错）",
                    ],
                    signal_name=sig.name,
                    message_name=msg.name,
                    input_value=str(sig_min - sig_factor),
                    priority="medium",
                )
            )

        elif method == TestMethod.NORMAL_RANGE:
            mid = (sig_min + sig_max) / 2
            cases.append(
                TestCase(
                    id=f"{prefix}_NR_MID",
                    name=f"{sig.name} 正常中间值",
                    category="正常范围测试",
                    method=method.value,
                    description=f"验证信号 {sig.name} 在正常中间值 ({mid:.2f}) 时的通信",
                    steps=[
                        f"设置 {msg.name}.{sig.name} = {mid:.2f} {sig_unit}",
                        f"以周期发送报文 {msg.name}",
                        "持续发送100个周期",
                    ],
                    expected_results=[
                        f"接收方正确解析 {sig.name} ≈ {mid:.2f} {sig_unit}",
                        "无丢帧或超时",
                    ],
                    signal_name=sig.name,
                    message_name=msg.name,
                    input_value=f"{mid:.2f}",
                    priority="medium",
                )
            )

        elif method == TestMethod.ERROR_INJECTION:
            cases.append(
                TestCase(
                    id=f"{prefix}_EI_INVALID",
                    name=f"{sig.name} 无效原始值",
                    category="错误注入测试",
                    method=method.value,
                    description=f"向 {sig.name} 注入超出物理范围的原始值",
                    steps=[
                        f"构造原始值使 {sig.name} 物理值 > {sig_max}",
                        f"发送报文 {msg.name} (0x{msg.id:03X})",
                    ],
                    expected_results=[
                        "接收方正确处理无效值（使用默认值或报警）",
                    ],
                    signal_name=sig.name,
                    message_name=msg.name,
                    priority="high",
                )
            )

        elif method == TestMethod.SIGNAL_TIMEOUT:
            cases.append(
                TestCase(
                    id=f"{prefix}_TIMEOUT",
                    name=f"{sig.name} 信号超时",
                    category="超时测试",
                    method=method.value,
                    description=f"验证 {msg.name} 报文停止发送后 {sig.name} 的超时处理",
                    steps=[
                        f"正常发送 {msg.name} 10秒",
                        f"停止发送 {msg.name}",
                        "等待超时时间",
                    ],
                    expected_results=[
                        f"接收方检测到 {sig.name} 超时",
                        "使用安全默认值或触发DTC",
                    ],
                    signal_name=sig.name,
                    message_name=msg.name,
                    priority="critical",
                )
            )

        return cases

    def generate_message_tests(self, methods: list[TestMethod] | None = None) -> int:
        """Generate message-level test cases (E2E, Counter, etc.)."""
        if self.current_dbc is None:
            return 0
        if methods is None:
            methods = [TestMethod.E2E_PROTECTION, TestMethod.COUNTER_VALIDATION]

        count = 0
        for msg in self.current_dbc.messages:
            for method in methods:
                cases = self._generate_for_message(msg, method)
                self.test_cases.extend(cases)
                count += len(cases)
        return count

    def _generate_for_message(self, msg: MessageDef, method: TestMethod) -> list[TestCase]:
        cases = []
        prefix = f"TC_{msg.name}"

        if method == TestMethod.E2E_PROTECTION:
            # E2E CRC 校验
            cases.append(
                TestCase(
                    id=f"{prefix}_E2E_CRC",
                    name=f"{msg.name} E2E CRC 校验",
                    category="E2E保护测试",
                    method=method.value,
                    description=f"验证报文 {msg.name} (0x{msg.id:03X}) 的 E2E CRC 计算与校验",
                    steps=[
                        f"正常发送报文 {msg.name}，CRC 字段按协议计算",
                        "接收方校验 CRC",
                        "修改 CRC 字段为错误值后发送",
                        "观察接收方是否检测到 CRC 错误",
                    ],
                    expected_results=[
                        "正常 CRC 时接收方正确接受报文",
                        "错误 CRC 时接收方拒绝报文或使用安全值",
                        "CRC 错误计数器递增",
                    ],
                    message_name=msg.name,
                    priority="critical",
                )
            )
            # E2E Counter 校验
            cases.append(
                TestCase(
                    id=f"{prefix}_E2E_CNT",
                    name=f"{msg.name} E2E 计数器校验",
                    category="E2E保护测试",
                    method=method.value,
                    description=f"验证报文 {msg.name} 的 E2E 计数器按协议递增",
                    steps=[
                        f"连续发送报文 {msg.name}，观察 E2E 计数器字段",
                        "确认计数器按 0→Max→0 循环递增",
                        "发送计数器跳变的报文（如 0→5）",
                        "观察接收方是否检测到计数器异常",
                    ],
                    expected_results=[
                        "计数器按协议定义的步长和范围递增",
                        "计数器跳变时接收方检测到错误",
                        "连续丢帧时接收方在超时后报错",
                    ],
                    message_name=msg.name,
                    priority="critical",
                )
            )
            # E2E 超时检测
            cases.append(
                TestCase(
                    id=f"{prefix}_E2E_TIMEOUT",
                    name=f"{msg.name} E2E 超时检测",
                    category="E2E保护测试",
                    method=method.value,
                    description=f"验证报文 {msg.name} 停止发送后 E2E 超时机制",
                    steps=[
                        f"正常发送 {msg.name} 5秒",
                        "停止发送",
                        "等待 E2E 超时时间（通常 2-5 个报文周期）",
                    ],
                    expected_results=[
                        "接收方在超时后检测到 E2E 错误",
                        "相关信号使用安全默认值",
                        "触发对应的 DTC（如有定义）",
                    ],
                    message_name=msg.name,
                    priority="high",
                )
            )
            # E2E 状态位校验
            cases.append(
                TestCase(
                    id=f"{prefix}_E2E_STATUS",
                    name=f"{msg.name} E2E 状态位校验",
                    category="E2E保护测试",
                    method=method.value,
                    description=f"验证报文 {msg.name} 的 E2E 状态指示位",
                    steps=[
                        f"正常发送 {msg.name}，读取 E2E 状态信号",
                        "注入 CRC 错误，读取 E2E 状态信号",
                        "注入计数器错误，读取 E2E 状态信号",
                    ],
                    expected_results=[
                        "正常时 E2E 状态 = OK",
                        "CRC 错误时 E2E 状态 = CRC_Error",
                        "计数器错误时 E2E 状态 = Counter_Error",
                    ],
                    message_name=msg.name,
                    priority="high",
                )
            )

        elif method == TestMethod.COUNTER_VALIDATION:
            # 计数器递增
            cases.append(
                TestCase(
                    id=f"{prefix}_CNT_INC",
                    name=f"{msg.name} 计数器递增",
                    category="计数器测试",
                    method=method.value,
                    description=f"验证报文 {msg.name} 的滚动计数器每周期递增",
                    steps=[
                        f"连续发送 20 个周期的 {msg.name} 报文",
                        "记录每帧的计数器字段值",
                        "验证相邻帧计数器差值 = 1",
                    ],
                    expected_results=[
                        "计数器每帧递增 1",
                        "无跳变或重复",
                    ],
                    message_name=msg.name,
                    priority="high",
                )
            )
            # 计数器回绕
            cases.append(
                TestCase(
                    id=f"{prefix}_CNT_WRAP",
                    name=f"{msg.name} 计数器回绕",
                    category="计数器测试",
                    method=method.value,
                    description=f"验证报文 {msg.name} 的计数器在最大值后正确回绕",
                    steps=[
                        f"持续发送 {msg.name} 直到计数器达到最大值",
                        "记录最大值和回绕后的值",
                    ],
                    expected_results=[
                        "计数器在达到最大值后回到 0",
                        "回绕过程无丢帧或异常",
                    ],
                    message_name=msg.name,
                    priority="high",
                )
            )
            # 计数器冻结检测
            cases.append(
                TestCase(
                    id=f"{prefix}_CNT_FREEZE",
                    name=f"{msg.name} 计数器冻结检测",
                    category="计数器测试",
                    method=method.value,
                    description="验证接收方检测到计数器冻结（连续相同值）",
                    steps=[
                        f"正常发送 {msg.name} 几个周期",
                        "连续发送计数器值相同的报文（冻结）",
                        "观察接收方反应",
                    ],
                    expected_results=[
                        "接收方检测到计数器冻结",
                        "触发超时或错误处理机制",
                    ],
                    message_name=msg.name,
                    priority="high",
                )
            )
            # 计数器跳变检测
            cases.append(
                TestCase(
                    id=f"{prefix}_CNT_JUMP",
                    name=f"{msg.name} 计数器跳变检测",
                    category="计数器测试",
                    method=method.value,
                    description="验证接收方检测到计数器跳变（非连续值）",
                    steps=[
                        f"正常发送 {msg.name}，计数器递增中",
                        "突然发送计数器跳变的报文（如从 3 直接到 8）",
                        "观察接收方是否检测到跳变",
                    ],
                    expected_results=[
                        "接收方检测到计数器不连续",
                        "报文被标记为无效或触发错误处理",
                    ],
                    message_name=msg.name,
                    priority="medium",
                )
            )

        return cases

    # ── Test case CRUD ─────────────────────────────────────────────────────

    def get_test_cases(self) -> list[TestCase]:
        return self.test_cases

    def get_cases_by_category(self) -> dict[str, list[TestCase]]:
        cats: dict[str, list[TestCase]] = {}
        for tc in self.test_cases:
            cats.setdefault(tc.category, []).append(tc)
        return cats

    def get_cases_by_message(self) -> dict[str, list[TestCase]]:
        msgs: dict[str, list[TestCase]] = {}
        for tc in self.test_cases:
            msgs.setdefault(tc.message_name or "未关联", []).append(tc)
        return msgs

    def get_coverage(self) -> dict:
        """Calculate signal coverage statistics."""
        if self.current_dbc is None:
            return {"total_signals": 0, "covered": 0, "coverage": 0}
        total_signals: set[str] = set()
        signal_index: dict[str, str] = {}  # signal_name → "msg.signal"
        for msg in self.current_dbc.messages:
            for sig in msg.signals:
                key = f"{msg.name}.{sig.name}"
                total_signals.add(key)
                signal_index[sig.name] = key
        covered_signals = set()
        for tc in self.test_cases:
            if tc.signal_name and tc.signal_name in signal_index:
                covered_signals.add(signal_index[tc.signal_name])
        total = len(total_signals)
        covered = len(covered_signals)
        return {
            "total_signals": total,
            "covered": covered,
            "coverage": (covered / total * 100) if total > 0 else 0,
            "total_cases": len(self.test_cases),
        }

    def add_test_case(self, tc: TestCase):
        self.test_cases.append(tc)

    def remove_test_case(self, tc_id: str):
        self.test_cases = [tc for tc in self.test_cases if tc.id != tc_id]

    def update_status(self, tc_id: str, status: str):
        for tc in self.test_cases:
            if tc.id == tc_id:
                tc.status = status
                return True
        return False

    def clear_cases(self):
        self.test_cases.clear()

    # ── Export ──────────────────────────────────────────────────────────────

    def export_json(self, output_path: Path) -> tuple[bool, list[str]]:
        try:
            data = {
                "test_cases": [
                    {
                        "id": tc.id,
                        "name": tc.name,
                        "category": tc.category,
                        "method": tc.method,
                        "description": tc.description,
                        "preconditions": tc.preconditions,
                        "steps": tc.steps,
                        "expected_results": tc.expected_results,
                        "signal_name": tc.signal_name,
                        "message_name": tc.message_name,
                        "input_value": tc.input_value,
                        "priority": tc.priority,
                        "status": tc.status,
                    }
                    for tc in self.test_cases
                ],
                "coverage": self.get_coverage(),
            }
            output_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
            return True, []
        except Exception as exc:
            return False, [str(exc)]

    def export_excel(self, output_path: Path) -> tuple[bool, list[str]]:
        try:
            from openpyxl import Workbook

            from core.utils.excel_utils import auto_width, write_header_row

            wb = Workbook()
            ws = wb.active
            ws.title = "测试用例"

            headers = [
                "用例ID",
                "名称",
                "分类",
                "方法",
                "描述",
                "前置条件",
                "测试步骤",
                "预期结果",
                "关联信号",
                "关联报文",
                "输入值",
                "优先级",
                "状态",
            ]
            write_header_row(ws, headers)

            for row, tc in enumerate(self.test_cases, 2):
                ws.cell(row=row, column=1, value=tc.id)
                ws.cell(row=row, column=2, value=tc.name)
                ws.cell(row=row, column=3, value=tc.category)
                ws.cell(row=row, column=4, value=tc.method)
                ws.cell(row=row, column=5, value=tc.description)
                ws.cell(row=row, column=6, value="\n".join(tc.preconditions))
                ws.cell(row=row, column=7, value="\n".join(tc.steps))
                ws.cell(row=row, column=8, value="\n".join(tc.expected_results))
                ws.cell(row=row, column=9, value=tc.signal_name)
                ws.cell(row=row, column=10, value=tc.message_name)
                ws.cell(row=row, column=11, value=tc.input_value)
                ws.cell(row=row, column=12, value=tc.priority)
                ws.cell(row=row, column=13, value=tc.status)

            auto_width(ws)
            ws.freeze_panes = "A2"
            wb.save(str(output_path))
            return True, []
        except (OSError, ValueError) as exc:
            return False, [str(exc)]
